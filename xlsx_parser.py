import os
import bs4
from bs4 import BeautifulSoup
import openpyxl


class Cursor:
    x = 1
    y = 1

MONTHS = {
    '01': "January",
    '02': "February",
    '03': "March",
    '04': "April",
    '05': "May",
    '06': "June",
    '07': "July",
    '08': "August",
    '09': "September",
    '10': "October",
    '11': "November",
    '12': "December"
}

TABLE_COORDINATES = {
    'DAYSTATS'    : 3,
    'HOURSTATS'   : 5,
    'TOPURLS'     : 6,
    'URLS_BY_KB'  : 7,
    'TOPENTRY'    : 8,
    'TOPEXIT'     : 9,
    'TOPSITES'    : 10,
    'SITES_BY_KB' : 11,
    'TOPREFS'     : 12,
    'TOPSEARCH'   : 13,
    'TOPAGENTS'   : 14,
    'TOPCTRYS'    : 16
    }


def extract_tables(html):
    '''
        Parse the html document and create a list of tuples where each tuple
        contains the table's name and its data.

        This function only works for the specific strucutre of the webalizer
        stat pages I had access to.

        The TABLE_COORDINATES objects containst each table title and the index
        of the paragraph where the table can be found when all paragraphs of
        the document are listed.
    '''
    soup = BeautifulSoup(html, 'html.parser')
    tables = []

    all_paragraphs = soup.center.find_all('p')

    for table_title in TABLE_COORDINATES:
        coord = TABLE_COORDINATES[table_title]
        tables.append((table_title, all_paragraphs[coord].find('table')))
    return tables


def get_sheet_by_name(wb, sheetname):
    if sheetname not in wb.sheetnames:
        if 'Sheet' in wb.sheetnames:
            sheet = wb['Sheet']
            sheet.title = sheetname
        else:
            sheet = wb.create_sheet(sheetname)

    return wb[sheetname]


def table_to_xlsx_sheet(raw_tables, sheet):
    '''
        This function will paste the table data into the sheet.

        It creates a cursor object pointing to the x, y position of a single
        cell at the sheet. The cursor will be moved as needed.
    '''
    cur = Cursor()
    for title, table in raw_tables:
        sheet.cell(column=cur.x, row=cur.y).value = title

        for tr in table.find_all('tr'):
            for Tcell in tr.find_all(['th', 'td']):

                #  Checking for the rowspan and colspan of the html objects we
                # can merge the corresponding cells of the sheet to fit the
                # same format.

                rowspan = 1
                colspan = 1
                if Tcell.has_attr('colspan'):
                    colspan = int(Tcell['colspan'])
                if Tcell.has_attr('rowspan'):
                    rowspan = int(Tcell['rowspan'])

                if sheet.cell(column=cur.x, row=cur.y).coordinate in sheet.merged_cells:
                    cur.x += 1

                if rowspan > 1 or colspan > 1:
                    sheet.merge_cells(start_row=cur.y, end_row=cur.y+rowspan-1,
                                         start_column=cur.x, end_column=cur.x+colspan-1)

                # After the merging we add the contents
                sheet.cell(column=cur.x, row=cur.y).value = Tcell.get_text()

                cur.x += colspan

            cur.y += 1
            cur.x = 1

        cur.y += 1

    return sheet


def execute(directory, xlsx_file):
    '''
        Create a xls workbook, read all downloaded stat pages and parse them
        into the spreadsheet.
    '''
    wb = openpyxl.Workbook()
    for file_ in os.listdir(directory):
        sheet_name = "{}-{}".format(MONTHS[file_[-7:-5]], file_[-11:-7])

        with open(os.path.join(directory, file_)) as f:
            html = f.read()
            f.close()

        tables = extract_tables(html)
        table_to_xlsx_sheet(tables, get_sheet_by_name(wb, sheet_name))

    wb.save("{}{}".format(xlsx_file, '.xlsx'))


if __name__ == '__main__':
    directory = input("Type in the directory where the pages are located:")
    xlsx_file = input("Type in the path to save the xlsx file:")

    if not os.path.isdir(directory):
        print("The given directory does not exist.")
        exit()

    execute(directory, xlsx_file)
